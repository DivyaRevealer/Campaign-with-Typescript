"""Sales order related API endpoints."""

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re
from typing import Annotated, Any, Dict, Iterable, Iterator, Mapping, Optional, Tuple
from zipfile import BadZipFile

from anyio import fail_after
try:
    from anyio.exceptions import TimeoutError as AnyIOTimeout
except ImportError:
    try:
        from anyio import TooSlowError as AnyIOTimeout
    except ImportError:
        AnyIOTimeout = TimeoutError
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel
from sqlalchemy import delete, exists, func, insert, select
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.audit import log_audit
from app.core.concurrency import run_in_thread_limited
from app.core.config import settings
from app.core.rate_limit import limiter
from app.core.db import get_session, repeatable_read_transaction
from app.core.db_errors import raise_on_lock_conflict
from app.core.db_retry import with_db_retry
from app.core.optimistic_lock import _ensure_expected_timestamp
from app.core.idempotency import (
    IdempotencyClaimState,
    bump_idempotency_heartbeat,
    claim_idempotency_key,
    complete_idempotency_key,
    require_idempotency_key,
)
from app.core.deps import get_current_user
from app.models.inv_currency import InvCurrencyMaster
from app.models.inv_excel_upload import InvExcelUpload
from app.models.inv_generic_sequence import InvGenericSequence
from app.models.inv_sales_order import InvSoDtl, InvSoHdr, InvSoSubDtl
from app.models.inv_user import InvUserMaster
from app.schemas.salesorder import (
    SalesOrderCancelOut,
    SalesOrderCancelPayload,
    SalesOrderItemOut,
    SalesOrderOut,
    SalesOrderPayload,
    SalesOrderUploadItemOut,
    SalesOrderUploadOut,
    SalesOrderItemPayload,
    SalesOrderHeaderPayload,
)

router = APIRouter(prefix="/sales-orders", tags=["sales-orders"])

COL_KEYS: Tuple[str, ...] = ("description", "part_no", "due_on", "qty", "rate", "per", "disc_pct")
DEFAULT_COLUMN_ALIASES: Dict[str, str] = {
    "description": "description",
    "descriptionofgoods": "description",
    "description_goods": "description",
    "itemdescription": "description",
    "goodsdescription": "description",
    "partno": "part_no",
    "partnumber": "part_no",
    "itemcode": "part_no",
    "itemno": "part_no",
    "productcode": "part_no",
    "sopartno": "part_no",
    "materialid": "part_no",
    "dueon": "due_on",
    "duedate": "due_on",
    "deliverydate": "due_on",
    "deliveryon": "due_on",
    "qty": "qty",
    "quantity": "qty",
    "quantityordered": "qty",
    "rate": "rate",
    "price": "rate",
    "unitprice": "rate",
    "per": "per",
    "unit": "per",
    "uom": "per",
    "disc": "disc_pct",
    "discpct": "disc_pct",
    "discount": "disc_pct",
    "discountpct": "disc_pct",
    "discountpercent": "disc_pct",
    "discpercent": "disc_pct",
}


def _normalise_key_token(value: Any) -> str:
    token = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]", "", token)


def _load_column_aliases() -> Dict[str, str]:
    """Load column aliases from the shared JSON config.

    Falls back to DEFAULT_COLUMN_ALIASES if the config file is missing or invalid.
    """

    config_root = Path(__file__).resolve().parents[4] / "config" / "excel_column_aliases.json"
    try:
        with config_root.open("r", encoding="utf-8") as fp:
            data = json.load(fp)
            alias_map: Dict[str, str] = {}
            for key, value in (data or {}).items():
                token = _normalise_key_token(key)
                if value in COL_KEYS:
                    alias_map[token] = value
            if alias_map:
                return alias_map
            logger.warning("excel_aliases_config_empty_fallback", path=str(config_root))
    except FileNotFoundError:
        logger.warning("excel_aliases_config_missing_fallback", path=str(config_root))
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception("excel_aliases_config_failed", path=str(config_root), exc=exc)
    fallback: Dict[str, str] = {}
    for key, value in DEFAULT_COLUMN_ALIASES.items():
        token = _normalise_key_token(key)
        if value in COL_KEYS:
            fallback[token] = value
    return fallback


COLUMN_ALIASES: Dict[str, str] = _load_column_aliases()

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
DEFAULT_EXCEL_EPOCH = datetime(1899, 12, 31)


class SalesOrderUploadJsonIn(BaseModel):
    template: str
    sheet_name: str
    rows: list[dict[str, Any]]


class ScanUploadOut(BaseModel):
    status: str
    detail: str | None = None


def _normalise_voucher_no(value: str | None) -> str:
    return (value or "").strip().upper()


def _load_workbook_from_bytes(raw_bytes: bytes):
    return load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)


def _save_workbook_to_buffer(workbook: Workbook, buffer: BytesIO) -> None:
    workbook.save(buffer)


def _build_sales_order_sequence_key(order_date: date) -> tuple[str, int]:
    year = order_date.year
    return f"sales_order:{year}", year


async def _reserve_sequence_number(session: AsyncSession, seq_name: str) -> int:
    """Reserve and return the next value for the provided sequence name."""

    max_attempts = 5
    attempts = 0
    while True:
        attempts += 1
        row = await session.scalar(
            select(InvGenericSequence)
            .where(InvGenericSequence.seq_name == seq_name)
            .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
        )
        if row:
            current = int(row.seq_no or 1)
            row.seq_no = current + 1
            await session.flush()
            return current
        try:
            await session.execute(
                insert(InvGenericSequence).values(seq_name=seq_name, seq_no=1)
            )
            await session.flush()
        except IntegrityError:
            # Another transaction inserted the sequence concurrently. Retry the loop
            if attempts >= max_attempts:
                logger.bind(seq_name=seq_name).warning("sequence_reservation_exhausted")
                raise HTTPException(
                    status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                    detail="Sequence allocation failed. Please retry.",
                )
            continue


async def _peek_sales_order_number(session: AsyncSession, order_date: date) -> str:
    """
    Return the next suggested sales order number without reserving it.

    This intentionally avoids taking locks so concurrent callers can all see the
    same suggestion while the real reservation happens later inside the create
    transaction. That means the preview can be stale, but it will never consume
    or block on the sequence record, keeping the critical write path free of
    extra contention.
    """
    seq_name, year = _build_sales_order_sequence_key(order_date)
    current = await session.scalar(
        select(InvGenericSequence.seq_no).where(InvGenericSequence.seq_name == seq_name)
    )
    next_value = int(current or 1)
    return f"SO-{year}-{next_value:06d}"


async def _generate_sales_order_number(
    session: AsyncSession, order_date: date
) -> str:
    seq_name, year = _build_sales_order_sequence_key(order_date)
    next_value = await _reserve_sequence_number(session, seq_name)
    return f"SO-{year}-{next_value:06d}"


async def _next_so_sno(session: AsyncSession, so_no: str | int) -> int:
    stmt = (
        select(InvSoDtl.so_sno)
        .where(InvSoDtl.so_no == str(so_no))
        .order_by(InvSoDtl.so_sno.desc())
        .limit(1)
        .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
    )
    max_sno = (await session.execute(stmt)).scalar_one_or_none() or 0
    return int(max_sno) + 1


TWO_PLACES = Decimal("0.01")
HUNDRED = Decimal("100")

SubDetailKey = tuple[str, str, str]  # (so_no, so_prod_name, so_part_no)


def _normalise_existing_part_no(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if value is None:
        return ""
    return str(value).strip()


def _require_part_number(raw: Any, line_no: int) -> str:
    if isinstance(raw, str):
        part_no = raw.strip()
    elif raw is None:
        part_no = ""
    else:
        part_no = str(raw).strip()
    if not part_no:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Part number is required for line item {line_no}.",
        )
    return part_no


def _require_due_date(due_on: Optional[date], line_no: int) -> date:
    if due_on is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Due date is required for line item {line_no}.",
        )
    return due_on


def _require_uom(raw: Optional[str], line_no: int) -> str:
    value = (raw or "").strip().upper()
    if not value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unit of measure is required for line item {line_no}.",
        )
    return value


async def _load_sub_detail_map(
    session: AsyncSession, so_no: str, *, for_update: bool = False
) -> dict[SubDetailKey, InvSoSubDtl]:
    """Fetch aggregated sales order details keyed by order, product, and part."""

    stmt = select(InvSoSubDtl).where(InvSoSubDtl.so_no == so_no)
    if for_update:
        stmt = stmt.with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
    result = await session.execute(stmt)
    rows = result.scalars().all()
    return {
        (row.so_no, row.so_prod_name, _normalise_existing_part_no(row.so_part_no)): row
        for row in rows
    }


def _resolve_sub_totals(
    sub_detail_map: Mapping[SubDetailKey, InvSoSubDtl] | None,
    so_no: str,
    prod_name: str,
    part_no: Optional[str],
) -> tuple[Decimal, Decimal, Decimal]:
    """Return production, delivery, and stock totals for a given key."""

    if not sub_detail_map:
        return (Decimal("0"), Decimal("0"), Decimal("0"))

    key = (so_no, prod_name, _normalise_existing_part_no(part_no))
    sub_detail = sub_detail_map.get(key)
    if not sub_detail:
        return (Decimal("0"), Decimal("0"), Decimal("0"))

    return (
        sub_detail.prod_qty or Decimal("0"),
        sub_detail.dely_qty or Decimal("0"),
        sub_detail.stk_qty or Decimal("0"),
    )


def _quantise(value: Decimal, scale: Decimal = TWO_PLACES) -> Decimal:
    """Round decimal values to the desired scale (defaults to 2 decimal places)."""

    return value.quantize(scale)


def _prepare_line_items(so_no: str, items: list[SalesOrderItemPayload]) -> list[InvSoDtl]:
    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one line item is required.",
        )

    prepared: list[InvSoDtl] = []
    for idx, item in enumerate(items, start=1):
        line_ref = int(item.line_no) if item.line_no else idx
        qty = item.qty
        rate = item.rate
        disc_pct = item.disc_pct or Decimal("0")

        if qty <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item quantity must be greater than zero.",
            )
        if rate <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item rate must be greater than zero.",
            )
        if disc_pct < 0 or disc_pct > HUNDRED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Discount percent must be between 0 and 100.",
            )

        part_no = _require_part_number(item.part_no, line_ref)
        due_on = _require_due_date(item.due_on, line_ref)
        uom = _require_uom(item.per, line_ref)

        line_total = qty * rate
        discount = (line_total * disc_pct) / HUNDRED
        amount = line_total - discount

        if amount <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item amount must be greater than zero after discount.",
            )

        prepared.append(
            InvSoDtl(
                so_no=so_no,
                so_sno=idx,
                so_prod_name=item.description.strip(),
                so_part_no=part_no,
                so_due_on=due_on,
                so_qty=_quantise(qty),
                so_rate=_quantise(rate),
                so_uom=uom,
                so_disc_per=_quantise(disc_pct) if disc_pct is not None else None,
                so_amount=_quantise(amount),
            )
        )

    return prepared


def _prepare_update_line_items(
    so_no: str,
    items: list[SalesOrderItemPayload],
    existing_snos: set[int],
    sub_detail_map: Mapping[SubDetailKey, InvSoSubDtl] | None,
) -> tuple[Dict[int, Dict[str, Any]], list[Dict[str, Any]]]:
    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one line item is required.",
        )

    existing_updates: Dict[int, Dict[str, Any]] = {}
    new_items: list[Dict[str, Any]] = []
    totals_by_key: Dict[SubDetailKey, Decimal] = {}

    for idx, item in enumerate(items, start=1):
        qty = item.qty
        rate = item.rate
        disc_pct = item.disc_pct or Decimal("0")

        if qty <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item quantity must be greater than zero.",
            )
        if rate <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item rate must be greater than zero.",
            )
        if disc_pct < 0 or disc_pct > HUNDRED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Discount percent must be between 0 and 100.",
            )

        description = item.description.strip()
        if not description:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item description is required.",
            )

        part_no = _require_part_number(item.part_no, idx)
        due_on = _require_due_date(item.due_on, idx)
        uom = _require_uom(item.per, idx)
        line_total = qty * rate
        discount = (line_total * disc_pct) / HUNDRED
        amount = line_total - discount

        if amount <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Line item amount must be greater than zero after discount.",
            )

        sanitised: Dict[str, Any] = {
            "so_prod_name": description,
            "so_part_no": part_no,
            "so_due_on": due_on,
            "so_qty": _quantise(qty),
            "so_rate": _quantise(rate),
            "so_uom": uom,
            "so_disc_per": _quantise(disc_pct) if disc_pct is not None else None,
            "so_amount": _quantise(amount),
        }

        key = (so_no, sanitised["so_prod_name"], sanitised["so_part_no"])
        running_total = totals_by_key.get(key, Decimal("0"))
        totals_by_key[key] = running_total + (sanitised["so_qty"] or Decimal("0"))

        maybe_existing = item.line_no
        if maybe_existing is not None and maybe_existing in existing_snos:
            existing_updates[int(maybe_existing)] = sanitised
        else:
            new_items.append(sanitised)

    keys_to_validate: set[SubDetailKey] = set(totals_by_key)
    if sub_detail_map:
        keys_to_validate.update(sub_detail_map.keys())

    for key in keys_to_validate:
        key_so_no, prod_name, part_no_value = key
        produced, delivered, stock = _resolve_sub_totals(
            sub_detail_map, key_so_no, prod_name, part_no_value
        )
        min_required = max(produced, delivered, stock)
        if min_required <= 0:
            continue

        total_qty = totals_by_key.get(key, Decimal("0"))
        if total_qty >= min_required:
            continue

        part_display = f" (Part No: {part_no_value})" if part_no_value else ""
        required_display = format(min_required, "f")
        produced_display = format(produced, "f")
        delivered_display = format(delivered, "f")
        stock_display = format(stock, "f")
        after_display = format(total_qty, "f")

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Cannot reduce total quantity for product "
                f"'{prod_name}'{part_display} below {required_display}. "
                f"(Produced {produced_display}, Delivered {delivered_display}, "
                f"Stock {stock_display}; after change = {after_display})."
            ),
        )

    return existing_updates, new_items


async def _sync_sales_order_subtotals(session: AsyncSession, so_no: str) -> None:
    """Ensure ``inv_so_sub_dtl`` reflects the latest aggregated quantities."""

    await session.execute(
        select(InvSoDtl.so_no).where(InvSoDtl.so_no == so_no).with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
    )

    totals_stmt = (
        select(
            InvSoDtl.so_no,
            InvSoDtl.so_prod_name,
            InvSoDtl.so_part_no,
            func.sum(InvSoDtl.so_qty).label("total_qty"),
        )
        .where(InvSoDtl.so_no == so_no)
        .group_by(InvSoDtl.so_no, InvSoDtl.so_prod_name, InvSoDtl.so_part_no)
    )
    totals = (await session.execute(totals_stmt)).all()

    existing_stmt = (
        select(InvSoSubDtl)
        .where(InvSoSubDtl.so_no == so_no)
        .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
    )
    existing_rows = (await session.execute(existing_stmt)).scalars().all()
    existing_map = {
        (row.so_no, row.so_prod_name, _normalise_existing_part_no(row.so_part_no)): row
        for row in existing_rows
    }

    seen: set[SubDetailKey] = set()

    for so_no_value, prod_name, part_no, total_qty in totals:
        normalised_part_no = _normalise_existing_part_no(part_no)
        key = (so_no_value, prod_name, normalised_part_no)
        seen.add(key)
        quantised_total = _quantise(total_qty or Decimal("0"))
        maybe_existing = existing_map.get(key)
        if maybe_existing:
            maybe_existing.so_qty = quantised_total
        else:
            session.add(
                InvSoSubDtl(
                    so_no=so_no_value,
                    so_prod_name=prod_name,
                    so_part_no=normalised_part_no,
                    so_qty=quantised_total,
                    prod_qty=_quantise(Decimal("0")),
                    dely_qty=_quantise(Decimal("0")),
                )
            )

    for key, existing in existing_map.items():
        if key not in seen:
            session.delete(existing)

    stale_stmt = (
        delete(InvSoSubDtl)
        .where(InvSoSubDtl.so_no == so_no)
        .where(
            ~exists()
            .where(InvSoDtl.so_no == InvSoSubDtl.so_no)
            .where(InvSoDtl.so_prod_name == InvSoSubDtl.so_prod_name)
            .where(func.trim(func.coalesce(InvSoDtl.so_part_no, "")) == func.trim(func.coalesce(InvSoSubDtl.so_part_no, "")))
            .correlate(InvSoSubDtl)
        )
    )
    await session.execute(stale_stmt)

    await session.flush()


def _serialise_sales_order(
    header: InvSoHdr,
    sub_detail_map: Mapping[SubDetailKey, InvSoSubDtl] | None = None,
) -> SalesOrderOut:
    items = sorted(header.items, key=lambda item: item.so_sno)
    header_out = {
        "so_voucher_no": header.so_no,
        "so_voucher_date": header.so_date,
        "order_date": header.so_date,
        "job_ref_no": header.job_ref_no,
        "client_po_no": header.client_po_no,
        "company_code": header.company_code,
        "company_name": header.company_name,
        "client_code": header.client_code,
        "client_name": header.client_name,
        "currency": header.currency_code,
        "so_status": header.so_status,
        "created_by": header.created_by,
        "created_at": header.created_at,
        "updated_by": header.updated_by,
        "updated_at": header.updated_at,
    }

    item_out = []
    for detail in items:
        prod_qty, dely_qty, stock_qty = _resolve_sub_totals(
            sub_detail_map, header.so_no, detail.so_prod_name, detail.so_part_no
        )
        item_out.append(
            SalesOrderItemOut(
                line_no=detail.so_sno,
                description=detail.so_prod_name,
                part_no=detail.so_part_no,
                due_on=detail.so_due_on,
                qty=float(detail.so_qty or Decimal("0")),
                rate=float(detail.so_rate or Decimal("0")),
                per=detail.so_uom,
                disc_pct=float(detail.so_disc_per or Decimal("0")),
                amount=float(detail.so_amount or Decimal("0")),
                prod_qty=float(prod_qty or Decimal("0")),
                dely_qty=float(dely_qty or Decimal("0")),
                stock_qty=float(stock_qty or Decimal("0")),
            )
        )

    return SalesOrderOut(header=header_out, items=item_out)


async def _load_sales_order_response(
    session: AsyncSession, so_no: str
) -> SalesOrderOut | None:
    header = await session.scalar(
        select(InvSoHdr)
        .options(selectinload(InvSoHdr.items))
        .where(InvSoHdr.so_no == so_no)
    )
    if not header:
        return None
    sub_detail_map = await _load_sub_detail_map(session, so_no)
    return _serialise_sales_order(header, sub_detail_map)


def _sanitise_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9.+-]", "", value.strip())
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    return None


def _decimal_to_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    quantised = value.quantize(Decimal("0.01"))
    text = format(quantised, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def _parse_due_date(value: Any, epoch: datetime) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
        except (TypeError, ValueError):
            return None
        return converted.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%b-%Y",
            "%d %b %Y",
            "%d.%m.%Y",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return None


def _is_blank_row(values: Iterable[Any]) -> bool:
    for cell in values:
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip():
                return False
        else:
            return False
    return True


def _build_header_map(values: Tuple[Any, ...]) -> Dict[int, str]:
    header_map: Dict[int, str] = {}
    for idx, cell in enumerate(values):
        alias = COLUMN_ALIASES.get(_normalise_key_token(cell))
        if alias:
            header_map[idx] = alias
    return header_map


def _build_record(values: Tuple[Any, ...], header_map: Dict[int, str] | None) -> Dict[str, Any]:
    record: Dict[str, Any] = {}
    if header_map:
        for idx, key in header_map.items():
            record[key] = values[idx] if idx < len(values) else None
    else:
        for idx, key in enumerate(COL_KEYS):
            record[key] = values[idx] if idx < len(values) else None
    return record


def _build_record_from_mapping(row: Mapping[str, Any]) -> Dict[str, Any]:
    record: Dict[str, Any] = {}
    for raw_key, value in row.items():
        alias = COLUMN_ALIASES.get(_normalise_key_token(raw_key))
        if alias in COL_KEYS:
            record[alias] = value
    for key in COL_KEYS:
        record.setdefault(key, None)
    return record


def _build_item(record: Dict[str, Any], epoch: datetime) -> SalesOrderUploadItemOut | None:
    description = _sanitise_text(record.get("description"))
    part_no = _sanitise_text(record.get("part_no"))
    due_on = _parse_due_date(record.get("due_on"), epoch)
    qty = _parse_decimal(record.get("qty"))
    rate = _parse_decimal(record.get("rate"))
    per = _sanitise_text(record.get("per")).upper()
    disc_pct = _parse_decimal(record.get("disc_pct"))

    if not any([description, part_no, due_on, qty, rate, per, disc_pct]):
        return None

    return SalesOrderUploadItemOut(
        description=description,
        part_no=part_no or None,
        due_on=due_on.isoformat() if due_on else None,
        qty=_decimal_to_string(qty),
        rate=_decimal_to_string(rate),
        per=per or None,
        disc_pct=_decimal_to_string(disc_pct),
    )


def _parse_json_rows(rows: Iterable[Mapping[str, Any]], *, epoch: datetime) -> list[SalesOrderUploadItemOut]:
    parsed: list[SalesOrderUploadItemOut] = []
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        record = _build_record_from_mapping(row)
        maybe_item = _build_item(record, epoch)
        if maybe_item:
            parsed.append(maybe_item)
    return parsed


def build_row_level_validation_errors(
    json_rows: list[Mapping[str, Any]],
    items: list[SalesOrderUploadItemOut],
) -> list[dict[str, Any]]:
    """Produce lightweight row-level validation errors for JSON uploads."""

    errors: list[dict[str, Any]] = []
    if not json_rows or not items:
        return errors

    for idx, row in enumerate(json_rows):
        if not isinstance(row, Mapping):
            continue
        record = _build_record_from_mapping(row)

        description = _sanitise_text(record.get("description"))
        part_no = _sanitise_text(record.get("part_no"))
        due_on_raw = record.get("due_on")
        qty = _parse_decimal(record.get("qty"))
        rate = _parse_decimal(record.get("rate"))
        per = _sanitise_text(record.get("per"))
        disc_pct = _parse_decimal(record.get("disc_pct"))

        if not any([description, part_no, due_on_raw, qty, rate, per, disc_pct]):
            continue

        if not description and any([part_no, due_on_raw, qty, rate, per, disc_pct]):
            errors.append({"row_index": idx, "message": "Description is required."})
            continue
        if not part_no and any([description, due_on_raw, qty, rate, per, disc_pct]):
            errors.append({"row_index": idx, "message": "Part number is required."})
            continue

        if due_on_raw is not None:
            due_on = _parse_due_date(due_on_raw, DEFAULT_EXCEL_EPOCH)
            if due_on_raw not in ("", None) and due_on is None:
                errors.append({"row_index": idx, "message": "Invalid due date."})
                continue

        if qty is not None and qty <= 0:
            errors.append({"row_index": idx, "message": "Quantity must be greater than zero."})
            continue

        if rate is not None and rate <= 0:
            errors.append({"row_index": idx, "message": "Rate must be greater than zero."})
            continue

        if disc_pct is not None and (disc_pct < 0 or disc_pct > HUNDRED):
            errors.append(
                {
                    "row_index": idx,
                    "message": "Discount % must be between 0 and 100.",
                }
            )

    return errors


def _parse_sheet(sheet, epoch: datetime) -> tuple[list[SalesOrderUploadItemOut], list[dict[str, Any]]]:
    """Parse a sheet and retain the raw row mappings for validation feedback."""

    parsed_rows: list[SalesOrderUploadItemOut] = []
    raw_row_mappings: list[dict[str, Any]] = []

    iterator: Iterator[Tuple[Any, ...]] = sheet.iter_rows(values_only=True)

    first_row: Tuple[Any, ...] | None = None
    for candidate in iterator:
        if candidate is None:
            continue
        if _is_blank_row(candidate):
            continue
        first_row = candidate
        break

    if first_row is None:
        return parsed_rows, raw_row_mappings

    header_map = _build_header_map(first_row)
    has_header = len(header_map) >= 3

    def _record_with_defaults(values: Tuple[Any, ...], header: Dict[int, str] | None) -> Dict[str, Any]:
        record = _build_record(values, header)
        for key in COL_KEYS:
            record.setdefault(key, None)
        return record

    if not has_header:
        record = _record_with_defaults(first_row, None)
        maybe_item = _build_item(record, epoch)
        if maybe_item:
            parsed_rows.append(maybe_item)
            raw_row_mappings.append(record)

    for values in iterator:
        if values is None or _is_blank_row(values):
            continue
        record = _record_with_defaults(values, header_map if has_header else None)
        maybe_item = _build_item(record, epoch)
        if maybe_item:
            parsed_rows.append(maybe_item)
            raw_row_mappings.append(record)

    return parsed_rows, raw_row_mappings


@router.get("/check")
async def check_sales_order(
    request: Request,
    so_voucher_no: str,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
):
    voucher = _normalise_voucher_no(so_voucher_no)
    if not voucher:
        return {"exists": False}

    stmt = select(func.count()).select_from(InvSoHdr).where(InvSoHdr.so_no == voucher)
    exists = (await session.execute(stmt)).scalar_one() > 0

    await log_audit(
        session,
        user.inv_user_code,
        "sales_order",
        voucher,
        "CHECK_EXISTS",
        details={"exists": exists},
        remote_addr=(request.client.host if request.client else None),
        independent_txn=True,
    )

    return {"exists": exists}


@router.get("/next-number")
async def get_next_sales_order_number(
    request: Request,
    order_date: date | None = None,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> dict[str, str]:
    target_date = order_date or date.today()

    try:
        voucher = await _peek_sales_order_number(session, target_date)

        await log_audit(
            session,
            user.inv_user_code,
            "sales_order",
            voucher,
            "SUGGEST_NUMBER",
            details={"suggested": voucher, "order_year": target_date.year},
            remote_addr=(request.client.host if request.client else None),
        )
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    return {"so_voucher_no": voucher}


@router.get("/{so_voucher_no}/export")
async def export_sales_order(
    so_voucher_no: str,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
):
    voucher = _normalise_voucher_no(so_voucher_no)
    stmt = (
        select(InvSoHdr)
        .options(selectinload(InvSoHdr.items))
        .where(InvSoHdr.so_no == voucher)
    )
    header = await session.scalar(stmt)
    if not header:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sales order not found.",
        )

    items = sorted(header.items or [], key=lambda item: item.so_sno)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Order"

    columns = [
        "Sales Order No",
        "Sales Order Date",
        "Job Ref No",
        "Company Code",
        "Company Name",
        "Client PO No",
        "Client Code",
        "Client Name",
        "Currency",
        "Product Name",
        "Part No",
        "Due On",
        "Quantity",
        "Rate",
        "UOM",
        "Discount %",
        "Amount",
    ]
    sheet.append(columns)

    so_date_display = header.so_date.strftime("%d-%m-%Y") if header.so_date else ""
    for item in items:
        due_on_display = item.so_due_on.strftime("%d-%m-%Y") if item.so_due_on else ""
        sheet.append(
            [
                header.so_no,
                so_date_display,
                header.job_ref_no,
                header.company_code,
                header.company_name,
                header.client_po_no,
                header.client_code,
                header.client_name,
                header.currency_code,
                item.so_prod_name,
                item.so_part_no,
                due_on_display,
                float(item.so_qty) if item.so_qty is not None else None,
                float(item.so_rate) if item.so_rate is not None else None,
                item.so_uom,
                float(item.so_disc_per) if item.so_disc_per is not None else 0.0,
                float(item.so_amount) if item.so_amount is not None else None,
            ]
        )

    buffer = BytesIO()
    try:
        with fail_after(settings.EXCEL_OP_TIMEOUT_SEC):
            await run_in_thread_limited(_save_workbook_to_buffer, workbook, buffer)
    except AnyIOTimeout as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel processing timed out. Please retry.",
            headers={"Retry-After": "2"},
        ) from exc
    buffer.seek(0)

    filename = f"SalesOrder_{header.so_no}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    await log_audit(
        session,
        user.inv_user_code,
        "sales_order",
        voucher,
        "EXPORT",
        details={"item_count": len(items)},
        remote_addr=(request.client.host if request.client else None),
        independent_txn=True,
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{so_voucher_no}", response_model=SalesOrderOut)
async def get_sales_order(
    so_voucher_no: str,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderOut:
    voucher = _normalise_voucher_no(so_voucher_no)
    stmt = (
        select(InvSoHdr)
        .options(selectinload(InvSoHdr.items))
        .where(InvSoHdr.so_no == voucher)
    )
    header = await session.scalar(stmt)
    if not header:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sales order not found.",
        )

    sub_detail_map = await _load_sub_detail_map(session, voucher)

    response = _serialise_sales_order(header, sub_detail_map)

    await log_audit(
        session,
        user.inv_user_code,
        "sales_order",
        voucher,
        "VIEW",
        details={"item_count": len(response.items)},
        remote_addr=(request.client.host if request.client else None),
        independent_txn=True,
    )

    return response


@router.post("", response_model=SalesOrderOut, status_code=status.HTTP_201_CREATED)
async def create_sales_order(
    payload: SalesOrderPayload,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderOut:
    header_payload: SalesOrderHeaderPayload = payload.header
    preferred_voucher = _normalise_voucher_no(header_payload.so_voucher_no)
    company_code = header_payload.company_code.upper()
    client_code = header_payload.client_code.upper()
    currency = header_payload.currency.upper()
    idempotency_key = require_idempotency_key(request)

    attempts = 3
    candidate = preferred_voucher
    last_error: IntegrityError | None = None

    async def _create_once() -> SalesOrderOut:
        nonlocal candidate
        async with repeatable_read_transaction(session):
            claim = await claim_idempotency_key(
                session,
                idempotency_key=idempotency_key,
                resource="sales_order",
            )
            if (
                claim.state == IdempotencyClaimState.REPLAY
                and claim.record
                and claim.record.resource_id
            ):
                existing = await _load_sales_order_response(
                    session, claim.record.resource_id
                )
                if existing:
                    return existing
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Original request completed but the sales order was not found.",
                )
            if claim.state == IdempotencyClaimState.IN_PROGRESS:
                retry_after = str(claim.retry_after or 1)
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Another request with this Idempotency-Key is still running. Please retry shortly.",
                    headers={"Retry-After": retry_after},
                )

            await bump_idempotency_heartbeat(
                session, idempotency_key=idempotency_key
            )

            currency_exists = await session.scalar(
                select(InvCurrencyMaster.currency_code).where(
                    InvCurrencyMaster.currency_code == currency
                )
            )
            if not currency_exists:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid currency code.",
                )

            order_date = (
                header_payload.so_voucher_date or header_payload.order_date or date.today()
            )

            if preferred_voucher:
                match = re.match(r"^SO-(\d{4})-", preferred_voucher or "")
                if match and int(match.group(1)) != int(order_date.year):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Voucher number prefix year must match Sales Order Date year.",
                    )

            if not candidate:
                candidate = await _generate_sales_order_number(session, order_date)

            job_ref_no = (header_payload.job_ref_no or "").strip()
            if not job_ref_no:
                job_ref_no = f"{client_code}-{candidate}"

            existing = await session.scalar(
                select(InvSoHdr.so_no)
                .where(InvSoHdr.so_no == candidate)
                .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
            )
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Sales order already exists.",
                )

            header = InvSoHdr(
                so_no=candidate,
                so_date=order_date,
                job_ref_no=job_ref_no,
                company_code=company_code,
                company_name=header_payload.company_name,
                client_po_no=header_payload.client_po_no.strip(),
                client_code=client_code,
                client_name=header_payload.client_name,
                currency_code=currency,
                created_by=user.inv_user_code,
            )
            header.items = _prepare_line_items(candidate, payload.items)
            session.add(header)
            await session.flush()

            await _sync_sales_order_subtotals(session, candidate)
            response = await _load_sales_order_response(session, candidate)
            if not response:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Failed to load saved sales order.",
                )

            await log_audit(
                session,
                user.inv_user_code,
                "sales_order",
                candidate,
                "CREATE",
                details={"item_count": len(response.items)},
                remote_addr=(request.client.host if request.client else None),
            )
            resource_id = header.so_no or response.so_voucher_no or candidate
            await complete_idempotency_key(
                session, idempotency_key=idempotency_key, resource_id=resource_id
            )
            return response

    for attempt in range(attempts):
        try:
            # Execute the unit of work with DB deadlock/lock-wait retries
            return await with_db_retry(session, _create_once)
        except IntegrityError as exc:
            last_error = exc
            if attempt == attempts - 1:
                raise
            # Regenerate the voucher on next attempt if there was a uniqueness collision
            candidate = ""
            continue
        except OperationalError as exc:
            raise_on_lock_conflict(exc)

    # Should be unreachable; keep for completeness
    if last_error is not None:
        raise last_error
    raise HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail="Unable to create sales order.",
    )


@router.put("/{so_voucher_no}", response_model=SalesOrderOut)
async def update_sales_order(
    so_voucher_no: str,
    payload: SalesOrderPayload,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderOut:
    voucher = _normalise_voucher_no(so_voucher_no)

    async def _update_once() -> SalesOrderOut:
        async with repeatable_read_transaction(session):
            header = await session.scalar(
                select(InvSoHdr)
                .options(selectinload(InvSoHdr.items))
                .where(InvSoHdr.so_no == voucher)
                .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
            )
            if not header:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Sales order not found.",
                )
            if header.so_status != "O":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Only open sales orders can be updated.",
                )

            _ensure_expected_timestamp(header.updated_at, payload.expected_updated_at)

            header_payload: SalesOrderHeaderPayload = payload.header
            payload_voucher = _normalise_voucher_no(header_payload.so_voucher_no)
            if payload_voucher and payload_voucher != voucher:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Voucher number cannot be changed.",
                )

            company_code = header_payload.company_code.upper()
            client_code = header_payload.client_code.upper()
            currency = header_payload.currency.upper()
            currency_exists = await session.scalar(
                select(InvCurrencyMaster.currency_code).where(
                    InvCurrencyMaster.currency_code == currency
                )
            )
            if not currency_exists:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid currency code.",
                )
            order_date = (
                header_payload.so_voucher_date
                or header_payload.order_date
                or header.so_date
            )

            job_ref_no = (header_payload.job_ref_no or "").strip()
            if not job_ref_no:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Job reference is required.",
                )

            sub_detail_map = await _load_sub_detail_map(
                session, voucher, for_update=True
            )

            so_no = voucher
            existing_rows = await session.execute(
                select(InvSoDtl)
                .where(InvSoDtl.so_no == so_no)
                .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
            )
            existing_by_sno = {row.so_sno: row for row in existing_rows.scalars()}

            existing_updates, new_items = _prepare_update_line_items(
                so_no,
                payload.items,
                set(existing_by_sno.keys()),
                sub_detail_map,
            )

            snos_to_delete = [
                sno for sno in existing_by_sno.keys() if sno not in existing_updates
            ]
            if snos_to_delete:
                await session.execute(
                    delete(InvSoDtl)
                    .where(InvSoDtl.so_no == so_no)
                    .where(InvSoDtl.so_sno.in_(snos_to_delete))
                )
                await session.flush()
                for sno in snos_to_delete:
                    existing_by_sno.pop(sno, None)

            for sno, values in existing_updates.items():
                detail = existing_by_sno.get(sno)
                if not detail:
                    new_items.append(values)
                    continue
                detail.so_prod_name = values["so_prod_name"]
                detail.so_part_no = values["so_part_no"]
                detail.so_due_on = values["so_due_on"]
                detail.so_qty = values["so_qty"]
                detail.so_rate = values["so_rate"]
                detail.so_uom = values["so_uom"]
                detail.so_disc_per = values["so_disc_per"]
                detail.so_amount = values["so_amount"]

            for values in new_items:
                new_sno = await _next_so_sno(session, so_no)
                session.add(
                    InvSoDtl(
                        so_no=so_no,
                        so_sno=new_sno,
                        so_prod_name=values["so_prod_name"],
                        so_part_no=values["so_part_no"],
                        so_due_on=values["so_due_on"],
                        so_qty=values["so_qty"],
                        so_rate=values["so_rate"],
                        so_uom=values["so_uom"],
                        so_disc_per=values["so_disc_per"],
                        so_amount=values["so_amount"],
                    )
                )

            header.so_date = order_date
            header.job_ref_no = job_ref_no
            header.client_po_no = header_payload.client_po_no.strip()
            header.company_code = company_code
            header.company_name = header_payload.company_name
            header.client_code = client_code
            header.client_name = header_payload.client_name
            header.currency_code = currency
            header.updated_by = user.inv_user_code
            header.updated_at = datetime.now()

            await session.flush()

            await _sync_sales_order_subtotals(session, so_no)

            sub_detail_map = await _load_sub_detail_map(session, voucher)

            fresh = await session.scalar(
                select(InvSoHdr)
                .options(selectinload(InvSoHdr.items))
                .where(InvSoHdr.so_no == voucher)
            )
            if not fresh:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Failed to load saved sales order.",
                )

            response = _serialise_sales_order(fresh, sub_detail_map)

            await log_audit(
                session,
                user.inv_user_code,
                "sales_order",
                voucher,
                "UPDATE",
                details={"item_count": len(response.items)},
                remote_addr=(request.client.host if request.client else None),
            )

            return response

    try:
        return await with_db_retry(session, _update_once)
    except OperationalError as exc:
        raise_on_lock_conflict(exc)


@router.post("/{so_voucher_no}/cancel", response_model=SalesOrderCancelOut)
async def cancel_sales_order(
    so_voucher_no: str,
    payload: SalesOrderCancelPayload,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderCancelOut:
    voucher = _normalise_voucher_no(so_voucher_no)

    async def _cancel_once() -> SalesOrderCancelOut:
        async with repeatable_read_transaction(session):
            header = await session.scalar(
                select(InvSoHdr)
                .options(selectinload(InvSoHdr.items))
                .where(InvSoHdr.so_no == voucher)
                .with_for_update(nowait=settings.DB_NOWAIT_LOCKS)
            )
            if not header:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Sales order not found.",
                )
            if header.so_status != "O":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Only open sales orders can be cancelled.",
                )

            _ensure_expected_timestamp(header.updated_at, payload.expected_updated_at)

            sub_detail_map = await _load_sub_detail_map(
                session, voucher, for_update=True
            )

            blocking_exists = any(
                (detail.prod_qty or Decimal("0")) > 0
                or (detail.dely_qty or Decimal("0")) > 0
                or (detail.stk_qty or Decimal("0")) > 0
                for detail in sub_detail_map.values()
            )
            if blocking_exists:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Cannot delete. Dependent transactions exist.",
                )

            header.so_status = "X"
            header.updated_by = user.inv_user_code
            header.updated_at = datetime.now()

            await session.flush()

            await log_audit(
                session,
                user.inv_user_code,
                "sales_order",
                voucher,
                "CANCEL",
                details={"item_count": len(header.items)},
                remote_addr=(request.client.host if request.client else None),
            )

            return SalesOrderCancelOut(
                so_voucher_no=voucher,
                status="cancelled",
                message="Sales order cancelled successfully.",
            )

    try:
        return await with_db_retry(session, _cancel_once)
    except OperationalError as exc:
        raise_on_lock_conflict(exc)


@router.post("/scan-upload", response_model=ScanUploadOut)
@limiter.limit(getattr(settings, "EXCEL_UPLOAD_RATE", "5/minute"))
async def scan_sales_order_upload(
    file: Annotated[UploadFile, File(...)],
    request: Request,
    user: InvUserMaster = Depends(get_current_user),
) -> ScanUploadOut:
    """Scan an uploaded Excel file without parsing it."""

    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file was uploaded.",
        )
    if filename.startswith("~$"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The Excel file appears to be open. Please close it and try again.",
        )

    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel workbooks (.xlsx, .xlsm) can be uploaded.",
        )

    try:
        raw_bytes = await file.read(settings.MAX_UPLOAD_BYTES + 1)
    finally:
        await file.close()

    if not raw_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file is empty.",
        )
    if len(raw_bytes) > settings.MAX_UPLOAD_BYTES:
        max_size_mb = settings.MAX_UPLOAD_BYTES / (1024 * 1024)
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Max allowed size is {max_size_mb:.0f} MB.",
        )

    if not settings.ENABLE_FILE_SCAN:
        return ScanUploadOut(status="clean", detail="File not scanned (scanning disabled).")

    from app.utils.security import scan_file_for_viruses

    is_clean = scan_file_for_viruses(raw_bytes)
    if not is_clean:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={
                "status": "infected",
                "detail": "The uploaded file failed security checks.",
            },
        )

    return ScanUploadOut(status="clean", detail="File passed security checks.")


@router.post("/upload-items-json", response_model=SalesOrderUploadOut)
@limiter.limit(getattr(settings, "EXCEL_UPLOAD_RATE", "5/minute"))
async def upload_sales_order_items_json(
    payload: SalesOrderUploadJsonIn,
    request: Request,
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderUploadOut:
    """Validate and parse pre-parsed JSON rows into line items."""

    # 1. Validate and normalise the template (file) name
    template_name = (payload.template or "").strip()
    base_name = Path(template_name).stem
    if not base_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file name.",
        )

    # 2. Look up authorised (file_name, sheet_name) combinations
    stmt = select(InvExcelUpload.file_name, InvExcelUpload.sheet_name).where(
        func.lower(InvExcelUpload.file_name) == base_name.lower()
    )
    authorised_entries = list((await session.execute(stmt)).all())
    if not authorised_entries:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The selected file is not registered for upload.",
        )

    authorised_lookup = {
        sheet.casefold(): (file_name, sheet) for file_name, sheet in authorised_entries
    }

    # 3. Enforce that the sheet provided by the UI is registered
    requested_sheet_name = (payload.sheet_name or "").strip()
    if not requested_sheet_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sheet name is required for this template.",
        )

    lookup_key = requested_sheet_name.casefold()
    if lookup_key not in authorised_lookup:
        # Example: DB has "SO Details", UI sends "SO Details1" -> this will fire
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The selected sheet is not registered for upload.",
        )

    registered_file_name, registered_sheet_name = authorised_lookup[lookup_key]

    # 4. Parse and validate rows
    items = _parse_json_rows(payload.rows or [], epoch=DEFAULT_EXCEL_EPOCH)
    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No line items were found in the selected sheet.",
        )

    row_errors = build_row_level_validation_errors(payload.rows or [], items)
    if row_errors:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={
                "detail": "Validation errors in uploaded Excel.",
                "row_errors": row_errors,
            },
        )

    # 5. Audit + response
    await log_audit(
        session,
        user.inv_user_code,
        "sales_order",
        None,
        "UPLOAD_ITEMS_JSON",
        details={
            "file_name": registered_file_name,
            "sheet_name": registered_sheet_name,
            "item_count": len(items),
            "original_template": template_name,
        },
        remote_addr=(request.client.host if request.client else None),
        independent_txn=True,
    )

    return SalesOrderUploadOut(
        file_name=registered_file_name,
        sheet_name=registered_sheet_name,
        items=items,
    )

@router.post("/upload-items", response_model=SalesOrderUploadOut)
@limiter.limit(getattr(settings, "EXCEL_UPLOAD_RATE", "5/minute"))
async def upload_sales_order_items(
    request: Request,
    file: Annotated[UploadFile, File(...)],
    session: AsyncSession = Depends(get_session),
    user: InvUserMaster = Depends(get_current_user),
) -> SalesOrderUploadOut:
    """Validate and parse a sales order Excel file into line items."""

    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file was uploaded.",
        )
    if filename.startswith("~$"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The Excel file appears to be open. Please close it and try again.",
        )

    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel workbooks (.xlsx, .xlsm) can be uploaded.",
        )

    base_name = Path(filename).stem
    if not base_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file name.",
        )

    stmt = select(InvExcelUpload.file_name, InvExcelUpload.sheet_name).where(
        func.lower(InvExcelUpload.file_name) == base_name.lower()
    )
    authorised_entries = list((await session.execute(stmt)).all())
    if not authorised_entries:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The selected file is not registered for upload.",
        )

    authorised_lookup = {
        sheet.casefold(): (file_name, sheet)
        for file_name, sheet in authorised_entries
    }

    try:
        raw_bytes = await file.read()
        if len(raw_bytes) > settings.MAX_UPLOAD_BYTES:
            max_size_mb = settings.MAX_UPLOAD_BYTES / (1024 * 1024)
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"File too large. Max allowed size is {max_size_mb:.0f} MB.",
            )
    except Exception as exc:  # pragma: no cover - defensive guard
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to read the uploaded file.",
        ) from exc
    finally:
        await file.close()

    if not raw_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file is empty.",
        )

    if settings.ENABLE_FILE_SCAN:
        from app.utils.security import scan_file_for_viruses

        if not scan_file_for_viruses(raw_bytes):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded file failed security checks.",
            )

    try:
        with fail_after(settings.EXCEL_OP_TIMEOUT_SEC):
            workbook = await run_in_thread_limited(
                _load_workbook_from_bytes, raw_bytes
            )
    except PermissionError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please close the Excel file before uploading and try again.",
        ) from exc
    except (InvalidFileException, BadZipFile) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file is not a valid Excel workbook.",
        ) from exc
    except AnyIOTimeout as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel processing timed out. Please retry.",
            headers={"Retry-After": "2"},
        ) from exc
    except Exception as exc:  # pragma: no cover - defensive guard
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The Excel file could not be opened.",
        ) from exc

    matched_sheet_name: str | None = None
    registered_file_name: str | None = None
    registered_sheet_name: str | None = None
    try:
        for sheet_name in workbook.sheetnames:
            key = sheet_name.casefold()
            if key in authorised_lookup:
                registered_file_name, registered_sheet_name = authorised_lookup[key]
                matched_sheet_name = sheet_name
                break

        if matched_sheet_name is None or registered_sheet_name is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded workbook does not contain the required sheet.",
            )

        sheet = workbook[matched_sheet_name]
        with fail_after(settings.EXCEL_OP_TIMEOUT_SEC):
            items, raw_rows = await run_in_thread_limited(
                _parse_sheet, sheet, workbook.epoch
            )
    except AnyIOTimeout as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel processing timed out. Please retry.",
            headers={"Retry-After": "2"},
        ) from exc
    finally:
        workbook.close()

    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No line items were found in the selected sheet.",
        )

    row_errors = build_row_level_validation_errors(raw_rows, items)
    if row_errors:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={
                "detail": "Validation errors in uploaded Excel.",
                "row_errors": row_errors,
            },
        )

    await log_audit(
        session,
        user.inv_user_code,
        "sales_order",
        None,
        "UPLOAD_ITEMS",
        details={
            "file_name": registered_file_name,
            "sheet_name": registered_sheet_name,
            "item_count": len(items),
            "original_filename": filename,
        },
        remote_addr=(request.client.host if request.client else None),
        independent_txn=True,
    )

    return SalesOrderUploadOut(
        file_name=registered_file_name or base_name,
        sheet_name=registered_sheet_name or matched_sheet_name,
        items=items,
    )